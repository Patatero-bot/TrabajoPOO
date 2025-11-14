from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from DAO.Conexion import ConexionBD

class ExportadorExcel:
    @staticmethod
    def generar_plantilla_empleados(conexion: ConexionBD, filename="plantilla_empleados.xlsx"):
        try:
            empleados = conexion.ejecutar(
                """
                SELECT p.nombre_completo, p.cargo, p.salario, p.email, p.fono, d.nombre as departamento, p.fecha_inicio
                FROM persona p LEFT JOIN departamento d ON p.iddepartamento = d.iddepartamento
                ORDER BY d.nombre, p.nombre_completo
                """
            ).fetchall()
            if not empleados:
                print("No hay empleados.")
                return False
            wb = Workbook()
            ws = wb.active
            ws.title = "Empleados"
            headers = ['Nombre Completo', 'Cargo', 'Salario', 'Email', 'Teléfono', 'Departamento', 'Fecha Inicio']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                c.alignment = Alignment(horizontal="center")
            for row, e in enumerate(empleados, 2):
                ws.cell(row=row, column=1, value=e['nombre_completo'])
                ws.cell(row=row, column=2, value=e['cargo'])
                ws.cell(row=row, column=3, value=e['salario'])
                ws.cell(row=row, column=4, value=e['email'])
                ws.cell(row=row, column=5, value=e['fono'])
                ws.cell(row=row, column=6, value=e['departamento'] or 'Sin asignar')
                ws.cell(row=row, column=7, value=e['fecha_inicio'].strftime('%d/%m/%Y') if e['fecha_inicio'] else '')
            for i, w in enumerate([30,20,15,25,15,20,15], 1):
                ws.column_dimensions[chr(64 + i)].width = w
            ws2 = wb.create_sheet("Resumen")
            ws2['A1'] = "RESUMEN DE EMPLEADOS"
            ws2['A1'].font = Font(bold=True, size=14)
            ws2['A3'] = "Total de empleados:"; ws2['B3'] = len(empleados)
            ws2['A4'] = "Fecha de generación:"; ws2['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
            wb.save(filename)
            print(f"Excel generado: {filename}")
            return True
        except Exception as e:
            print(f"Error Excel: {e}")
            return False

    @staticmethod
    def generar_plantilla_proyectos(conexion: ConexionBD, filename="plantilla_proyectos.xlsx"):
        try:
            proyectos = conexion.ejecutar(
                """
                SELECT p.nombre, p.descripcion, d.nombre as departamento, COUNT(ip.idpersona) as empleados_asignados
                FROM proyecto p LEFT JOIN departamento d ON p.iddepartamento = d.iddepartamento
                LEFT JOIN inscripcion_proyecto ip ON p.idproyecto = ip.idproyecto
                GROUP BY p.idproyecto ORDER BY d.nombre, p.nombre
                """
            ).fetchall()
            if not proyectos:
                print("No hay proyectos.")
                return False
            wb = Workbook()
            ws = wb.active
            ws.title = "Proyectos"
            headers = ['Proyecto', 'Descripción', 'Departamento', 'Empleados Asignados']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                c.alignment = Alignment(horizontal="center")
            for row, p in enumerate(proyectos, 2):
                ws.cell(row=row, column=1, value=p['nombre'])
                ws.cell(row=row, column=2, value=p['descripcion'])
                ws.cell(row=row, column=3, value=p['departamento'] or 'Sin asignar')
                ws.cell(row=row, column=4, value=p['empleados_asignados'])
            for i, w in enumerate([25, 35, 20, 20], 1):
                ws.column_dimensions[chr(64 + i)].width = w
            wb.save(filename)
            print(f"Excel proyectos generado: {filename}")
            return True
        except Exception as e:
            print(f"Error proyectos Excel: {e}")
            return False